"""
doc_engine.py - High-performance parallel file generation engine.

Generates multiple Excel files in parallel using ProcessPoolExecutor.
Each file gets unique DocumentNo values across the entire batch —
ranges are pre-computed so workers never contend for shared state.

Usage:
    import doc_engine
    summary = doc_engine.run(
        input_file="template.xlsx",
        output_folder="output/",
        file_count=20,
        record_counts=[10, 100, 1000],
        progress_callback=lambda done, total, msg: print(msg),
    )
"""

import os
import re
import time
import openpyxl
from openpyxl import Workbook
from concurrent.futures import ProcessPoolExecutor, as_completed


# ---------------------------------------------------------------------------
# Document Number Helpers
# ---------------------------------------------------------------------------

def _extract_doc_number_parts(doc_number: str) -> tuple:
    """Extract prefix, numeric value, and digit width from a document number.

    Examples:
        'Gstr101'    -> ('Gstr', 101, 3)
        'TPFEWB0001' -> ('TPFEWB', 1, 4)
    """
    match = re.search(r'(\d+)$', str(doc_number))
    if not match:
        raise ValueError(f"Document number '{doc_number}' does not end with a number")

    num_str = match.group(1)
    prefix = str(doc_number)[:match.start()]
    return prefix, int(num_str), len(num_str)


def _format_doc_number(prefix: str, number: int, min_width: int) -> str:
    """Format a document number preserving at least *min_width* digits.

    Allows natural growth beyond min_width:
        ('Gstr', 106, 3)    -> 'Gstr106'
        ('Gstr', 1000, 3)   -> 'Gstr1000'
        ('TPFEWB', 42, 4)   -> 'TPFEWB0042'
    """
    return f"{prefix}{number:0{min_width}d}"


# ---------------------------------------------------------------------------
# Worker (runs in a separate process)
# ---------------------------------------------------------------------------

def _write_file_task(
    header: list,
    template_data: list,
    doc_col_idx: int,
    doc_prefix: str,
    doc_min_width: int,
    start_number: int,
    record_count: int,
    output_path: str,
) -> dict:
    """Write a single Excel file with *record_count* rows and unique DocumentNo values.

    Uses openpyxl **write-only** mode so rows are streamed to disk —
    this keeps memory flat even for 200 000+ row files.
    """
    try:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet()

        # Write header row
        ws.append(header)

        template_len = len(template_data)

        # Write data rows — cycle through template, stamp unique DocumentNo
        for i in range(record_count):
            row = list(template_data[i % template_len])
            row[doc_col_idx] = _format_doc_number(
                doc_prefix, start_number + i, doc_min_width
            )
            ws.append(row)

        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        wb.save(output_path)

        return {
            "status": "success",
            "output_path": output_path,
            "record_count": record_count,
        }
    except Exception as e:
        return {
            "status": "error",
            "output_path": output_path,
            "record_count": record_count,
            "error": str(e),
        }


# ---------------------------------------------------------------------------
# Template Reader
# ---------------------------------------------------------------------------

def _read_template(input_file: str, ref_column_name: str = "DocumentNo") -> tuple:
    """Read the template Excel file (first sheet) and return everything workers need.

    Returns:
        header          – list of column header values
        template_data   – list[list] of non-empty data rows
        doc_col_idx     – 0-based index of the DocumentNo column
        doc_prefix      – e.g. 'Gstr'
        doc_min_width   – minimum digit width to preserve
        last_doc_num    – last numeric value found (e.g. 105)
    """
    wb = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
    ws = wb.active  # Always the first / default sheet

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        raise ValueError("Template file must have at least a header row and one data row")

    header = list(rows[0])

    # Locate the DocumentNo column (case-insensitive partial match)
    doc_col_idx = None
    for idx, col_name in enumerate(header):
        if col_name and ref_column_name in str(col_name):
            doc_col_idx = idx
            break

    if doc_col_idx is None:
        raise ValueError(
            f"Could not find a '{ref_column_name}' column in the template header.\n"
            f"Headers found: {header}"
        )

    # Collect non-empty data rows and track last DocumentNo value
    template_data = []
    last_doc_value = None
    for row in rows[1:]:
        if any(cell is not None for cell in row):
            template_data.append(list(row))
            if row[doc_col_idx] is not None:
                last_doc_value = row[doc_col_idx]

    if not template_data:
        raise ValueError("Template file has no data rows")
    if last_doc_value is None:
        raise ValueError("No DocumentNo values found in the template data")

    doc_prefix, last_doc_num, doc_min_width = _extract_doc_number_parts(str(last_doc_value))

    return header, template_data, doc_col_idx, doc_prefix, doc_min_width, last_doc_num


# ---------------------------------------------------------------------------
# Public API — orchestrator
# ---------------------------------------------------------------------------

def run(
    input_file: str,
    output_folder: str,
    file_count: int,
    record_counts: list,
    ref_column_name: str = "DocumentNo",
    progress_callback=None,
    max_workers: int = None,
) -> dict:
    """Generate Excel files in parallel.

    Parameters
    ----------
    input_file : str
        Path to the template .xlsx file.
    output_folder : str
        Base directory for output.  Sub-folders are created automatically
        for each record count, named ``<file_count>x<record_count>/``
        (e.g. ``20x50/``, ``20x1000/``).
    file_count : int
        Number of files to create **per** record count.
    record_counts : list[int]
        e.g. ``[10, 100, 1000, 10000]``
    progress_callback : callable, optional
        ``callback(completed: int, total: int, message: str)``
    max_workers : int, optional
        Defaults to ``min(cpu_count, total_tasks, 8)``.

    Returns
    -------
    dict
        Summary with timing, counts, and error information.
    """
    start_time = time.time()

    # ── Phase 1: Read template ────────────────────────────────────────────
    if progress_callback:
        progress_callback(0, 0, "Reading template file...")

    (header, template_data, doc_col_idx,
     doc_prefix, doc_min_width, last_doc_num) = _read_template(input_file, ref_column_name)

    template_info = (
        f"Template: {len(template_data)} rows | "
        f"DocumentNo column: '{header[doc_col_idx]}' | "
        f"Last value: {_format_doc_number(doc_prefix, last_doc_num, doc_min_width)}"
    )
    if progress_callback:
        progress_callback(0, 0, template_info)

    # ── Phase 2: Build task list with pre-computed doc-number ranges ──────
    tasks = []
    current_start = last_doc_num + 1  # begin right after the last existing number

    # Derive a clean base name from the template file for output naming
    template_basename = os.path.splitext(os.path.basename(input_file))[0]

    for rec_count in record_counts:
        # Folder: <output_folder>/<file_count>x<record_count>  e.g. "20x50"
        subfolder = os.path.join(output_folder, f"{file_count}x{rec_count}")

        for file_idx in range(file_count):
            # File: <template>_<files>x<records>_part01.xlsx
            filename = f"{template_basename}_{file_count}x{rec_count}_part{file_idx + 1:02d}.xlsx"
            output_path = os.path.join(subfolder, filename)
            tasks.append({
                "header": header,
                "template_data": template_data,
                "doc_col_idx": doc_col_idx,
                "doc_prefix": doc_prefix,
                "doc_min_width": doc_min_width,
                "start_number": current_start,
                "record_count": rec_count,
                "output_path": output_path,
            })
            current_start += rec_count  # reserve the entire range for this task

    total_tasks = len(tasks)
    total_records = sum(t["record_count"] for t in tasks)

    plan_msg = (
        f"Plan: {total_tasks} files | "
        f"{total_records:,} total records | "
        f"Record counts: {record_counts} x {file_count} files each"
    )
    if progress_callback:
        progress_callback(0, total_tasks, plan_msg)

    # ── Phase 3: Execute in parallel ─────────────────────────────────────
    if max_workers is None:
        max_workers = min(os.cpu_count() or 4, total_tasks, 8)

    completed = 0
    errors = []

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        future_to_task = {
            executor.submit(_write_file_task, **task): task
            for task in tasks
        }

        for future in as_completed(future_to_task):
            result = future.result()
            completed += 1

            if result["status"] == "error":
                errors.append(result)

            if progress_callback:
                rel_path = os.path.relpath(result["output_path"], output_folder)
                if result["status"] == "success":
                    msg = (
                        f"[{completed}/{total_tasks}] [OK] {rel_path} "
                        f"({result['record_count']:,} records)"
                    )
                else:
                    msg = (
                        f"[{completed}/{total_tasks}] [FAIL] {rel_path} "
                        f"ERROR: {result.get('error', 'unknown')}"
                    )
                progress_callback(completed, total_tasks, msg)

    # ── Phase 4: Summary ─────────────────────────────────────────────────
    elapsed = time.time() - start_time

    summary = {
        "total_files": total_tasks,
        "successful_files": total_tasks - len(errors),
        "total_records": total_records,
        "errors": len(errors),
        "error_details": errors,
        "elapsed_seconds": round(elapsed, 2),
        "workers_used": max_workers,
        "record_counts": record_counts,
        "file_count_per_record_count": file_count,
    }

    if progress_callback:
        status = f"with {len(errors)} error(s)" if errors else "successfully"
        progress_callback(
            total_tasks, total_tasks,
            f"=== Completed {status}! "
            f"{total_tasks} files | {total_records:,} records | "
            f"{elapsed:.1f}s | {max_workers} workers ==="
        )

    return summary
