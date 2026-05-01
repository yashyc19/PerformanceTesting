import openpyxl
import os
import re
from time import time
from icecream import ic

# Import the Config class from config.py
from config import Config

# Global variable to store the last document number
last_doc_number = None

def increment_document_number(doc_number, increment=1):
    '''Increment the document number by the given increment value'''
    # Use regex to extract the ending numbers from the document number
    match = re.search(r'(\d+)$', doc_number)
    if not match:
        raise ValueError(f"Document number {doc_number} does not end with a number")
    
    number = int(match.group(1))
    new_number = number + increment
    
    # Replace the old number with the new incremented number
    new_doc_number = re.sub(r'\d+$', f"{new_number:04d}", doc_number)
    return new_doc_number

def get_document_number_column(sheet):
    '''Find the column index of the document number'''
    for cell in sheet[1]:
        if cell.value and Config.REFERENCE_COLUMN in cell.value:
            return cell.column

    raise ValueError(f"Reference column '{Config.REFERENCE_COLUMN}' not found in the sheet")

def process_file(inputfile, outputfile):
    '''Process the file'''
    global last_doc_number
    
    # SETUP
    # Load the excel file as a workbook for in-place editing
    workbook = openpyxl.load_workbook(inputfile)
    
    # Work on the default sheet
    sheet = workbook.active
    
    # Check number of records and compare with the config RECORD_COUNT
    # If RECORD_COUNT is none, use the number of records in the sheet
    # If RECORD_COUNT is a number n, use n records
    record_count = Config.RECORD_COUNT if Config.RECORD_COUNT else sheet.max_row
    record_count = int(record_count) + 1    # Add 1 to include the header row

    # Find the column index of the document number
    doc_number_column = get_document_number_column(sheet)
    # get the last document number from the last row of the sheet
    if last_doc_number is None:
        last_doc_number = sheet.cell(row=sheet.max_row, column=doc_number_column).value
    ic(last_doc_number)
    
    # PROCESSING
    # Now if the record_count is more than the number of records in the sheet,
    # we need to add more records by copying the existing records
    if record_count > sheet.max_row:
        ic('checkpoint: add extra rows')
        original_data = list(sheet.iter_rows(min_row=2, values_only=True))  # Skip header row
        original_row_count = len(original_data)
        
        for i in range(record_count - sheet.max_row):
            row_data = original_data[i % original_row_count]

            # Increment the document number
            row_data = list(row_data)
            row_data[doc_number_column - 1] = increment_document_number(last_doc_number)
            last_doc_number = row_data[doc_number_column - 1]
            sheet.append(row_data)  # Append the new row to the sheet

    # If the record_count is less than the number of records in the sheet,
    # we need to remove the extra records
    else:
        ic('checkpoint: delete extra rows')
        for i in range(sheet.max_row - record_count):
            sheet.delete_rows(sheet.max_row)

    # Save the workbook
    workbook.save(outputfile)
    ic(outputfile)


def main():
    '''Main function'''
    
    try:
        # Get the input file path
        input_file = os.path.join(Config.DOWNLOADS, f'{Config.FILE_NAME}.xlsx')
        output_file = os.path.join(Config.DOWNLOADS, f'{Config.OUTPUT_FILE_NAME}_{Config.FILE_NAME}')
        
        # Check if the input file exists
        if not os.path.exists(input_file):
            ic(f"File not found: {input_file}")
            raise FileNotFoundError(f"File not found: {input_file}")
        
        # start time of the process
        start_time = time()

        # Process the files
        for iteration in range(Config.ITERATIONS):
            process_file(input_file, f'{output_file}_{iteration+1}.xlsx')

        # end time of the process
        end_time = time()

        ic('Run status')
        ic(Config.ITERATIONS)
        ic(Config.RECORD_COUNT)
        ic(end_time - start_time)

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()